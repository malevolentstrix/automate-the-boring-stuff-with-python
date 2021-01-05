# "sheetinverter.py" of Chapter 12 of the book
# By JITHIN JOHN
# TODO
import openpyxl
import os
import sys
wb = sys.argv[1]
wb = os.path.abspath(wb)
workbook = openpyxl.load_workbook(wb)
active = workbook.get_active_sheet()
if active.max_row <= 0 or active.max_column <= 0:
    print('Error')
if not active:
    print('Error')
inverted shap = workbook.create_sheet("Invertedsheet")
for row in range(1, active.max_row):
    for col in range(1, active.max_column):
        invert.cell(row=col, column=row).value = active.cell(
        row=row, column=col).value
workbook.save(wb)
workbook.close()
